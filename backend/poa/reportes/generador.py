"""
Generador de reportes en PDF y Excel para el sistema POA.
Adaptado al modelo Obra existente.
"""

import os
import tempfile
from datetime import datetime
from io import BytesIO
from typing import List, Dict, Any

# ReportLab para PDFs
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfgen import canvas

# openpyxl para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ConfigReporte:
    """Configuración para generación de reportes"""
    
    def __init__(
        self,
        nombre: str = "Reporte POA",
        tipo_reporte: str = "ejecutivo",
        periodo: str = "mensual",
        fecha_corte: str = None,
        formato_salida: str = "pdf",
        incluir_graficos: bool = True,
        incluir_anexos: bool = False
    ):
        self.nombre = nombre
        self.tipo_reporte = tipo_reporte
        self.periodo = periodo
        self.fecha_corte = fecha_corte or datetime.now().strftime('%Y-%m-%d')
        self.formato_salida = formato_salida
        self.incluir_graficos = incluir_graficos
        self.incluir_anexos = incluir_anexos


class GeneradorReportes:
    """Clase principal para generar reportes en múltiples formatos"""
    
    TIPOS_REPORTE = {
        'ejecutivo': 'Reporte Ejecutivo',
        'cartera': 'Cartera de Proyectos',
        'presupuesto': 'Análisis Presupuestal',
        'riesgos': 'Análisis de Riesgos',
        'territorial': 'Distribución Territorial',
        'avance': 'Avance de Obras'
    }
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Configura estilos personalizados para PDFs"""
        # Título principal
        self.styles.add(ParagraphStyle(
            name='TituloReporte',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#9F2241'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        # Subtítulo
        self.styles.add(ParagraphStyle(
            name='SubtituloReporte',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#333333'),
            spaceAfter=8,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        ))
        
        # Información general
        self.styles.add(ParagraphStyle(
            name='InfoGeneral',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#666666'),
            spaceAfter=6
        ))
    
    # ==================== GENERACIÓN PDF ====================
    
    def generar_pdf_reporte(self, obras: List[Any], config: ConfigReporte, estadisticas: Dict) -> str:
        """
        Genera reporte en formato PDF.
        
        Args:
            obras: Lista de objetos Obra del modelo Django
            config: Configuración del reporte
            estadisticas: Diccionario con datos agregados
            
        Returns:
            str: Ruta del archivo PDF generado
        """
        # Crear archivo temporal
        temp_file = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
        pdf_path = temp_file.name
        
        # Configurar documento
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=letter,
            rightMargin=50,
            leftMargin=50,
            topMargin=50,
            bottomMargin=50
        )
        
        # Contenido del reporte
        story = []
        
        # Encabezado
        story.extend(self._crear_encabezado_pdf(config, estadisticas))
        
        # Contenido según tipo de reporte
        if config.tipo_reporte == 'ejecutivo':
            story.extend(self._crear_contenido_ejecutivo_pdf(obras, estadisticas))
        elif config.tipo_reporte == 'cartera':
            story.extend(self._crear_contenido_cartera_pdf(obras, estadisticas))
        elif config.tipo_reporte == 'presupuesto':
            story.extend(self._crear_contenido_presupuesto_pdf(obras, estadisticas))
        elif config.tipo_reporte == 'riesgos':
            story.extend(self._crear_contenido_riesgos_pdf(obras, estadisticas))
        elif config.tipo_reporte == 'territorial':
            story.extend(self._crear_contenido_territorial_pdf(obras, estadisticas))
        else:
            story.extend(self._crear_contenido_generico_pdf(obras, estadisticas))
        
        # Construir PDF
        doc.build(story)
        
        return pdf_path
    
    def _crear_encabezado_pdf(self, config: ConfigReporte, estadisticas: Dict) -> List:
        """Crea el encabezado del reporte PDF"""
        elementos = []
        
        # Título
        titulo = self.TIPOS_REPORTE.get(config.tipo_reporte, 'Reporte POA')
        elementos.append(Paragraph(titulo, self.styles['TituloReporte']))
        elementos.append(Spacer(1, 0.2*inch))
        
        # Información general
        info_texto = f"""
        <b>Fecha de corte:</b> {config.fecha_corte}<br/>
        <b>Período:</b> {config.periodo.capitalize()}<br/>
        <b>Total de proyectos:</b> {estadisticas.get('total_proyectos', 0)}<br/>
        <b>Presupuesto total:</b> ${estadisticas.get('presupuesto_total', 0):,.2f}<br/>
        <b>Generado:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}
        """
        elementos.append(Paragraph(info_texto, self.styles['InfoGeneral']))
        elementos.append(Spacer(1, 0.3*inch))
        
        return elementos
    
    def _crear_contenido_ejecutivo_pdf(self, obras: List[Any], estadisticas: Dict) -> List:
        """Contenido para reporte ejecutivo"""
        elementos = []
        
        # Resumen ejecutivo
        elementos.append(Paragraph('Resumen Ejecutivo', self.styles['SubtituloReporte']))
        
        # KPIs principales
        kpis_data = [
            ['Indicador', 'Valor'],
            ['Total Proyectos', str(estadisticas.get('total_proyectos', 0))],
            ['Presupuesto Total', f"${estadisticas.get('presupuesto_total', 0):,.2f}"],
            ['Avance Promedio', f"{estadisticas.get('avance_promedio', 0):.1f}%"],
            ['Beneficiarios', f"{estadisticas.get('total_beneficiarios', 0):,}"]
        ]
        
        tabla_kpis = Table(kpis_data, colWidths=[3*inch, 3*inch])
        tabla_kpis.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9F2241')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elementos.append(tabla_kpis)
        elementos.append(Spacer(1, 0.3*inch))
        
        # Top proyectos
        if obras:
            elementos.append(Paragraph('Top 10 Proyectos por Presupuesto', self.styles['SubtituloReporte']))
            
            top_obras = sorted(obras, key=lambda x: x.presupuesto_modificado or 0, reverse=True)[:10]
            
            tabla_data = [['Programa', 'Área', 'Presupuesto', 'Avance']]
            for obra in top_obras:
                tabla_data.append([
                    str(obra.programa)[:40] if obra.programa else 'N/A',
                    str(obra.area_responsable)[:30] if obra.area_responsable else 'N/A',
                    f"${obra.presupuesto_modificado:,.0f}" if obra.presupuesto_modificado else '$0',
                    f"{obra.avance_fisico_pct:.1f}%" if obra.avance_fisico_pct else '0%'
                ])
            
            tabla_proyectos = Table(tabla_data, colWidths=[2.5*inch, 1.8*inch, 1.2*inch, 0.8*inch])
            tabla_proyectos.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7F1D3A')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            elementos.append(tabla_proyectos)
        
        return elementos
    
    def _crear_contenido_cartera_pdf(self, obras: List[Any], estadisticas: Dict) -> List:
        """Contenido para cartera de proyectos"""
        elementos = []
        
        elementos.append(Paragraph('Cartera Completa de Proyectos', self.styles['SubtituloReporte']))
        
        if obras:
            # Agrupar por estado
            por_estado = estadisticas.get('por_estado', {})
            
            if por_estado:
                elementos.append(Paragraph('Distribución por Estado', self.styles['Heading3']))
                
                estado_data = [['Estado', 'Cantidad', 'Porcentaje']]
                total = sum(por_estado.values())
                
                for estado, cantidad in por_estado.items():
                    porcentaje = (cantidad / total * 100) if total > 0 else 0
                    estado_data.append([
                        estado or 'Sin estado',
                        str(cantidad),
                        f"{porcentaje:.1f}%"
                    ])
                
                tabla_estados = Table(estado_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
                tabla_estados.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9F2241')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey)
                ]))
                
                elementos.append(tabla_estados)
        
        return elementos
    
    def _crear_contenido_presupuesto_pdf(self, obras: List[Any], estadisticas: Dict) -> List:
        """Contenido para análisis presupuestal"""
        elementos = []
        
        elementos.append(Paragraph('Análisis Presupuestal', self.styles['SubtituloReporte']))
        
        # Resumen presupuestal
        presupuesto_data = [
            ['Concepto', 'Monto'],
            ['Presupuesto Modificado', f"${estadisticas.get('presupuesto_total', 0):,.2f}"],
            ['Anteproyecto Total', f"${estadisticas.get('anteproyecto_total', 0):,.2f}"],
            ['Ejecutado (estimado)', f"${estadisticas.get('presupuesto_ejecutado', 0):,.2f}"],
        ]
        
        tabla_presupuesto = Table(presupuesto_data, colWidths=[3*inch, 3*inch])
        tabla_presupuesto.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9F2241')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        
        elementos.append(tabla_presupuesto)
        
        return elementos
    
    def _crear_contenido_riesgos_pdf(self, obras: List[Any], estadisticas: Dict) -> List:
        """Contenido para análisis de riesgos"""
        elementos = []
        
        elementos.append(Paragraph('Análisis de Riesgos', self.styles['SubtituloReporte']))
        
        # Proyectos de alto riesgo
        alto_riesgo = [o for o in obras if (o.riesgo_nivel or 0) >= 4]
        
        riesgo_data = [
            ['Nivel de Riesgo', 'Cantidad'],
            ['Alto (4-5)', str(len(alto_riesgo))],
            ['Medio (3)', str(len([o for o in obras if (o.riesgo_nivel or 0) == 3]))],
            ['Bajo (1-2)', str(len([o for o in obras if (o.riesgo_nivel or 0) <= 2]))]
        ]
        
        tabla_riesgos = Table(riesgo_data, colWidths=[3*inch, 2*inch])
        tabla_riesgos.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9F2241')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        
        elementos.append(tabla_riesgos)
        
        if alto_riesgo:
            elementos.append(Spacer(1, 0.2*inch))
            elementos.append(Paragraph('Proyectos de Alto Riesgo', self.styles['Heading3']))
            
            for obra in alto_riesgo[:5]:
                texto = f"• {obra.programa or 'Sin nombre'} - Riesgo nivel: {obra.riesgo_nivel}"
                elementos.append(Paragraph(texto, self.styles['Normal']))
        
        return elementos
    
    def _crear_contenido_territorial_pdf(self, obras: List[Any], estadisticas: Dict) -> List:
        """Contenido para distribución territorial"""
        elementos = []
        
        elementos.append(Paragraph('Distribución Territorial', self.styles['SubtituloReporte']))
        
        # Análisis por alcaldía
        por_alcaldia = estadisticas.get('por_alcaldia', {})
        
        if por_alcaldia:
            alcaldia_data = [['Alcaldía', 'Proyectos', 'Presupuesto']]
            
            for alcaldia, datos in sorted(por_alcaldia.items(), 
                                         key=lambda x: x[1].get('presupuesto', 0), 
                                         reverse=True)[:10]:
                alcaldia_data.append([
                    alcaldia or 'No especificada',
                    str(datos.get('cantidad', 0)),
                    f"${datos.get('presupuesto', 0):,.0f}"
                ])
            
            tabla_alcaldias = Table(alcaldia_data, colWidths=[2.5*inch, 1.5*inch, 2*inch])
            tabla_alcaldias.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9F2241')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            
            elementos.append(tabla_alcaldias)
        
        return elementos
    
    def _crear_contenido_generico_pdf(self, obras: List[Any], estadisticas: Dict) -> List:
        """Contenido genérico para otros tipos de reporte"""
        elementos = []
        
        elementos.append(Paragraph('Resumen General', self.styles['SubtituloReporte']))
        
        resumen_texto = f"""
        Este reporte contiene información general sobre {estadisticas.get('total_proyectos', 0)} proyectos
        del Programa Operativo Anual con un presupuesto total de ${estadisticas.get('presupuesto_total', 0):,.2f}.
        """
        
        elementos.append(Paragraph(resumen_texto, self.styles['Normal']))
        
        return elementos
    
    # ==================== GENERACIÓN EXCEL ====================
    
    def generar_excel_reporte(self, obras: List[Any], config: ConfigReporte, estadisticas: Dict) -> str:
        """
        Genera reporte en formato Excel.
        
        Args:
            obras: Lista de objetos Obra del modelo Django
            config: Configuración del reporte
            estadisticas: Diccionario con datos agregados
            
        Returns:
            str: Ruta del archivo Excel generado
        """
        # Crear archivo temporal
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        excel_path = temp_file.name
        
        wb = Workbook()
        
        # Hoja 1: Resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        self._crear_hoja_resumen_excel(ws_resumen, config, estadisticas)
        
        # Hoja 2: Datos detallados
        ws_datos = wb.create_sheet("Datos")
        self._crear_hoja_datos_excel(ws_datos, obras)
        
        # Hoja 3: Análisis
        ws_analisis = wb.create_sheet("Análisis")
        self._crear_hoja_analisis_excel(ws_analisis, estadisticas)
        
        # Guardar
        wb.save(excel_path)
        
        return excel_path
    
    def _crear_hoja_resumen_excel(self, ws, config: ConfigReporte, estadisticas: Dict):
        """Crea la hoja de resumen en Excel"""
        # Estilos
        header_fill = PatternFill(start_color="9F2241", end_color="9F2241", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=14)
        
        # Título
        ws['A1'] = self.TIPOS_REPORTE.get(config.tipo_reporte, 'Reporte POA')
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # Información general
        ws['A3'] = 'Fecha de corte:'
        ws['B3'] = config.fecha_corte
        ws['A4'] = 'Período:'
        ws['B4'] = config.periodo
        ws['A5'] = 'Generado:'
        ws['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        # KPIs
        ws['A7'] = 'Indicador'
        ws['B7'] = 'Valor'
        ws['A7'].fill = header_fill
        ws['A7'].font = header_font
        ws['B7'].fill = header_fill
        ws['B7'].font = header_font
        
        kpis = [
            ('Total Proyectos', estadisticas.get('total_proyectos', 0)),
            ('Presupuesto Total', f"${estadisticas.get('presupuesto_total', 0):,.2f}"),
            ('Avance Promedio', f"{estadisticas.get('avance_promedio', 0):.1f}%"),
            ('Beneficiarios', f"{estadisticas.get('total_beneficiarios', 0):,}")
        ]
        
        row = 8
        for indicador, valor in kpis:
            ws[f'A{row}'] = indicador
            ws[f'B{row}'] = valor
            row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _crear_hoja_datos_excel(self, ws, obras: List[Any]):
        """Crea la hoja de datos detallados en Excel"""
        # Encabezados
        headers = [
            'ID', 'Programa', 'Área Responsable', 'Presupuesto Modificado',
            'Avance Físico', 'Avance Financiero', 'Estado', 'Alcaldías',
            'Beneficiarios', 'Fecha Inicio', 'Fecha Término'
        ]
        
        header_fill = PatternFill(start_color="7F1D3A", end_color="7F1D3A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Datos
        for row, obra in enumerate(obras, 2):
            ws.cell(row, 1, obra.id_excel or '')
            ws.cell(row, 2, str(obra.programa)[:50] if obra.programa else '')
            ws.cell(row, 3, obra.area_responsable or '')
            ws.cell(row, 4, obra.presupuesto_modificado or 0)
            ws.cell(row, 5, obra.avance_fisico_pct or 0)
            ws.cell(row, 6, obra.avance_financiero_pct or 0)
            ws.cell(row, 7, obra.estatus_general or '')
            ws.cell(row, 8, obra.alcaldias or '')
            ws.cell(row, 9, obra.beneficiarios_num or 0)
            ws.cell(row, 10, obra.fecha_inicio_prog.strftime('%Y-%m-%d') if obra.fecha_inicio_prog else '')
            ws.cell(row, 11, obra.fecha_termino_prog.strftime('%Y-%m-%d') if obra.fecha_termino_prog else '')
        
        # Ajustar anchos
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _crear_hoja_analisis_excel(self, ws, estadisticas: Dict):
        """Crea la hoja de análisis en Excel"""
        header_fill = PatternFill(start_color="9F2241", end_color="9F2241", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        ws['A1'] = 'Análisis Estadístico'
        ws['A1'].font = Font(bold=True, size=14)
        
        # Distribución por estado
        ws['A3'] = 'Distribución por Estado'
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        
        por_estado = estadisticas.get('por_estado', {})
        row = 4
        for estado, cantidad in por_estado.items():
            ws[f'A{row}'] = estado or 'Sin estado'
            ws[f'B{row}'] = cantidad
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
